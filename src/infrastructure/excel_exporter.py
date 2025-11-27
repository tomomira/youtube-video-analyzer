"""
Excel エクスポート機能

動画情報をExcelファイル(.xlsx)に出力する
"""
import os
from typing import List
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from src.domain.models import VideoInfo
from src.utils.logger import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """
    Excelファイルにエクスポートするクラス

    動画情報をExcelファイル(.xlsx)に出力する
    """

    def __init__(self):
        """初期化"""
        logger.info("ExcelExporterを初期化しました")

    def export(self, videos: List[VideoInfo], file_path: str) -> None:
        """
        動画情報をExcelファイルにエクスポート

        Args:
            videos: エクスポートする動画情報リスト
            file_path: 出力先ファイルパス

        Raises:
            ValueError: videosが空の場合
            IOError: ファイル書き込みに失敗した場合
        """
        if not videos:
            raise ValueError("エクスポートする動画が指定されていません")

        logger.info(f"Excelエクスポート開始: {len(videos)}件の動画を {file_path} に出力")

        try:
            # ワークブック作成
            wb = Workbook()
            ws = wb.active
            ws.title = "YouTube動画リスト"

            # ヘッダー設定
            self._create_header(ws)

            # データ行を追加
            for idx, video in enumerate(videos, start=2):
                self._add_video_row(ws, idx, video)

            # カラム幅を自動調整
            self._auto_adjust_column_width(ws)

            # ファイル保存
            wb.save(file_path)

            logger.info(f"Excelエクスポート完了: {file_path}")

        except Exception as e:
            logger.error(f"Excelエクスポートエラー: {e}", exc_info=True)
            raise IOError(f"Excelファイルの書き込みに失敗しました: {e}")

    def _create_header(self, ws) -> None:
        """
        ヘッダー行を作成

        Args:
            ws: ワークシート
        """
        headers = [
            "No",
            "タイトル",
            "URL",
            "チャンネル名",
            "チャンネルID",
            "再生回数",
            "いいね数",
            "コメント数",
            "公開日",
            "動画の長さ",
            "種類",
            "概要",
            "タグ",
            "サムネイルURL"
        ]

        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _add_video_row(self, ws, row_idx: int, video: VideoInfo) -> None:
        """
        動画情報の行を追加

        Args:
            ws: ワークシート
            row_idx: 行インデックス
            video: 動画情報
        """
        # No
        ws.cell(row=row_idx, column=1).value = row_idx - 1

        # タイトル
        ws.cell(row=row_idx, column=2).value = video.title

        # URL（ハイパーリンク）
        url_cell = ws.cell(row=row_idx, column=3)
        url_cell.value = video.url
        url_cell.hyperlink = video.url
        url_cell.font = Font(color="0563C1", underline="single")

        # チャンネル名
        ws.cell(row=row_idx, column=4).value = video.channel_name

        # チャンネルID
        ws.cell(row=row_idx, column=5).value = video.channel_id

        # 再生回数
        views_cell = ws.cell(row=row_idx, column=6)
        views_cell.value = video.view_count
        views_cell.number_format = '#,##0'

        # いいね数
        likes_cell = ws.cell(row=row_idx, column=7)
        likes_cell.value = video.like_count
        likes_cell.number_format = '#,##0'

        # コメント数
        comments_cell = ws.cell(row=row_idx, column=8)
        comments_cell.value = video.comment_count
        comments_cell.number_format = '#,##0'

        # 公開日（タイムゾーン情報を削除）
        published_cell = ws.cell(row=row_idx, column=9)
        # Excelはタイムゾーン付きdatetimeをサポートしないため、タイムゾーンを削除
        published_at_naive = video.published_at.replace(tzinfo=None) if video.published_at.tzinfo else video.published_at
        published_cell.value = published_at_naive
        published_cell.number_format = 'yyyy-mm-dd'

        # 動画の長さ
        ws.cell(row=row_idx, column=10).value = video.duration_formatted

        # 種類
        video_type = "ショート" if video.is_short else "通常"
        ws.cell(row=row_idx, column=11).value = video_type

        # 概要
        ws.cell(row=row_idx, column=12).value = video.description

        # タグ（カンマ区切り）
        tags_str = ", ".join(video.tags) if video.tags else ""
        ws.cell(row=row_idx, column=13).value = tags_str

        # サムネイルURL（ハイパーリンク）
        if video.thumbnail_url:
            thumb_cell = ws.cell(row=row_idx, column=14)
            thumb_cell.value = video.thumbnail_url
            thumb_cell.hyperlink = video.thumbnail_url
            thumb_cell.font = Font(color="0563C1", underline="single")

    def _auto_adjust_column_width(self, ws) -> None:
        """
        カラム幅を自動調整

        Args:
            ws: ワークシート
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    if cell.value:
                        # セル値の長さを計算（日本語は2文字分としてカウント）
                        cell_value = str(cell.value)
                        length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                        max_length = max(max_length, length)
                except:
                    pass

            # 最大幅を制限（20〜80文字）
            adjusted_width = min(max(max_length + 2, 20), 80)
            ws.column_dimensions[column_letter].width = adjusted_width
